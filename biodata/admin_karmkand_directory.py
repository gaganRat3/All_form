from django.contrib import admin
from django.utils.html import format_html
from .models_karmkand_directory import GlobalKarmkandDirectoryEntry, KarmkandiMaharajDetails, LaghuRudraYajmanRegistration, ShivMandirShivalayInfo

@admin.register(GlobalKarmkandDirectoryEntry)
class GlobalKarmkandDirectoryEntryAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'dob', 'location', 'phone1', 'phone2',
        'brahman_activities', 'experience_years', 'other_skills',
        'service_level', 'employment_status',
        'photo_preview', 'visiting_card', 'terms_agreed', 'submitted_at'
    )
    readonly_fields = ('photo_preview', 'submitted_at')
    search_fields = ('name', 'location', 'phone1', 'phone2')
    list_filter = ('service_level', 'employment_status', 'terms_agreed')

    actions = ['export_as_excel', 'download_images_zip']

    def photo_preview(self, obj):
        from django.utils.html import format_html
        if obj.photo:
            filename = obj.photo.name.split('/')[-1]
            return format_html('<a href="{}" target="_blank">{}</a>', obj.photo.url, filename)
        return "-"
    photo_preview.allow_tags = True
    photo_preview.short_description = "Photo Preview"

    def export_as_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Karmkand Directory"
        headers = [
            'Name', 'DOB', 'Location', 'Phone1', 'Phone2', 'Brahman Activities',
            'Experience Years', 'Other Skills', 'Service Level', 'Employment Status',
            'Photo', 'Visiting Card', 'Terms Agreed', 'Submitted At'
        ]
        ws.append(headers)
        for obj in queryset:
            ws.append([
                obj.name, obj.dob, obj.location, obj.phone1, obj.phone2,
                obj.brahman_activities, obj.experience_years, obj.other_skills,
                obj.service_level, obj.employment_status,
                obj.photo.url if obj.photo else '',
                obj.visiting_card.url if obj.visiting_card else '',
                obj.terms_agreed, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=global_karmkand_directory.xlsx'
        wb.save(response)
        return response
    export_as_excel.short_description = "Export selected as Excel"

    def download_images_zip(self, request, queryset):
        import io
        import zipfile
        from django.http import HttpResponse
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for obj in queryset:
                if obj.photo:
                    photo_path = obj.photo.path
                    zip_file.write(photo_path, f"photos/{obj.photo.name.split('/')[-1]}")
                if obj.visiting_card:
                    card_path = obj.visiting_card.path
                    zip_file.write(card_path, f"visiting_cards/{obj.visiting_card.name.split('/')[-1]}")
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type="application/zip")
        response['Content-Disposition'] = 'attachment; filename=karmkand_images.zip'
        return response
    download_images_zip.short_description = "Download images as ZIP"


@admin.register(KarmkandiMaharajDetails)
class KarmkandiMaharajDetailsAdmin(admin.ModelAdmin):
    list_display = ('serial_number', 'maharaj_name', 'mobile_number', 'date_of_birth', 'laghurudra_experience', 'residence_city', 'submitted_at')
    search_fields = ('maharaj_name', 'mobile_number', 'residence_city')
    list_filter = ('residence_city', 'submitted_at')
    readonly_fields = ('submitted_at',)
    ordering = ['-submitted_at']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('submitted_at')

    def serial_number(self, obj):
        request = getattr(self, 'admin_view_request', None)
        if request is None:
            return '-'
        queryset = self.get_queryset(request)
        pk_list = list(queryset.values_list('pk', flat=True))
        try:
            return pk_list.index(obj.pk) + 1
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial No.'
    serial_number.admin_order_field = None

    def get_changelist_instance(self, request):
        self.admin_view_request = request
        return super().get_changelist_instance(request)


@admin.register(LaghuRudraYajmanRegistration)
class LaghuRudraYajmanRegistrationAdmin(admin.ModelAdmin):
    list_display = ('payment_status_badge', 'serial_number', 'registered_by', 'registered_mobile', 'husband_name', 'wife_name', 'city', 'contact_number', 'full_address_preview', 'submitted_at')
    search_fields = ('registered_by', 'registered_mobile', 'husband_name', 'wife_name', 'city', 'full_address')
    list_filter = ('city', 'payment_status', 'submitted_at')
    readonly_fields = ('submitted_at',)
    ordering = ['-submitted_at']
    actions = ['mark_payment_success', 'mark_payment_pending', 'export_as_excel']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('submitted_at')

    def serial_number(self, obj):
        request = getattr(self, 'admin_view_request', None)
        if request is None:
            return '-'
        queryset = self.get_queryset(request)
        pk_list = list(queryset.values_list('pk', flat=True))
        try:
            return pk_list.index(obj.pk) + 1
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial No.'
    serial_number.admin_order_field = None

    def get_changelist_instance(self, request):
        self.admin_view_request = request
        return super().get_changelist_instance(request)

    def payment_status_badge(self, obj):
        status = obj.payment_status or 'pending'
        colors = {
            'pending': '#f59e0b',
            'paid': '#16a34a',
            'partial': '#3b82f6',
            'unpaid': '#ef4444',
        }
        label = obj.get_payment_status_display() if hasattr(obj, 'get_payment_status_display') else status.title()
        color = colors.get(status, '#64748b')
        return format_html(
            '<span style="display:inline-block;padding:6px 12px;border-radius:999px;background:{};color:#fff;font-weight:700;font-size:12px;min-width:120px;text-align:center;">{}</span>',
            color,
            label,
        )
    payment_status_badge.short_description = 'Payment Status'

    @admin.action(description='Mark selected as Payment Success')
    def mark_payment_success(self, request, queryset):
        queryset.update(payment_status='paid')

    @admin.action(description='Mark selected as Pending')
    def mark_payment_pending(self, request, queryset):
        queryset.update(payment_status='pending')

    @admin.action(description='Export selected registrations to Excel')
    def export_as_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.utils import get_column_letter

        ordered_queryset = self.get_queryset(request).filter(pk__in=queryset.values_list('pk', flat=True))

        headers = [
            'Payment Status',
            'Serial No.',
            'Registered By',
            'Registered Mobile',
            'Husband Name',
            'Wife Name',
            'City',
            'Contact Number',
            'Full Address',
            'Submitted At',
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Laghu Rudra Yajman'
        ws.append(headers)

        for obj in ordered_queryset:
            ws.append([
                obj.get_payment_status_display() if hasattr(obj, 'get_payment_status_display') else (obj.payment_status or 'Pending'),
                self.serial_number(obj),
                obj.registered_by,
                obj.registered_mobile,
                obj.husband_name,
                obj.wife_name,
                obj.city,
                obj.contact_number,
                obj.full_address,
                obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else '',
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    value = str(cell.value) if cell.value is not None else ''
                    max_length = max(max_length, len(value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=laghu_rudra_yajman_registrations.xlsx'
        wb.save(response)
        return response

    def full_address_preview(self, obj):
        return obj.full_address[:80] + ('...' if len(obj.full_address) > 80 else '')
    full_address_preview.short_description = 'Full Address'


@admin.register(ShivMandirShivalayInfo)
class ShivMandirShivalayInfoAdmin(admin.ModelAdmin):
    list_display = ('serial_number', 'temple_name', 'priest_president_name', 'priest_president_phone', 'city', 'form_filled_by', 'form_filler_phone', 'submitted_at')
    search_fields = ('temple_name', 'priest_president_name', 'city', 'form_filled_by')
    list_filter = ('city', 'submitted_at')
    readonly_fields = ('submitted_at',)
    ordering = ['-submitted_at']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('submitted_at')

    def serial_number(self, obj):
        request = getattr(self, 'admin_view_request', None)
        if request is None:
            return '-'
        queryset = self.get_queryset(request)
        pk_list = list(queryset.values_list('pk', flat=True))
        try:
            return pk_list.index(obj.pk) + 1
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial No.'
    serial_number.admin_order_field = None

    def get_changelist_instance(self, request):
        self.admin_view_request = request
        return super().get_changelist_instance(request)
