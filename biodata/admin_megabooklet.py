from django.contrib import admin
from django.utils.html import format_html
from .models import MegaBookletCorrectionRequest
from .models import CandidateBiodata
import zipfile
import os
from django.http import HttpResponse
from io import BytesIO
import requests

@admin.register(MegaBookletCorrectionRequest)
class MegaBookletCorrectionRequestAdmin(admin.ModelAdmin):
    list_display = ('serial_number', 'candidate_name', 'dob', 'city', 'whatsapp_number', 'booklet_serial', 'booklet_name', 'correction_description', 'photo_upload_preview', 'submission_time')
    search_fields = ('candidate_name', 'city', 'whatsapp_number', 'booklet_serial', 'booklet_name')
    list_filter = ('city', 'submission_time')
    readonly_fields = ('submission_time', 'dob', 'correction_description', 'photo_upload_preview')

    actions = ['export_selected_to_excel', 'download_selected_images']

    def photo_upload_preview(self, obj):
        if obj.photo_upload:
            return format_html('<img src="{}" style="max-height: 100px; max-width: 100px;" />', obj.photo_upload.url)
        return "-"
    photo_upload_preview.short_description = 'Photo Upload Preview'

    import zipfile
    import os
    from django.http import HttpResponse
    from io import BytesIO

    import requests

    def download_selected_images(self, request, queryset):
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for obj in queryset:
                if obj.photo_upload:
                    # Format filename as serialnumber_name_dob_whatsappnumber
                    serial = self.serial_number(obj)
                    name = obj.candidate_name.replace(' ', '_')
                    dob = obj.dob.replace(' ', '_') if obj.dob else 'unknown_dob'
                    whatsapp = obj.whatsapp_number.replace(' ', '_')
                    from urllib.parse import urlparse
                    url_path = urlparse(obj.photo_upload.url).path
                    ext = os.path.splitext(url_path)[1]
                    if not ext:
                        ext = '.jpg'  # default extension
                    filename = f"{serial}_{name}_{dob}_{whatsapp}{ext}"
                    # Download file content from URL
                    try:
                        response = requests.get(obj.photo_upload.url)
                        if response.status_code == 200:
                            zip_file.writestr(filename, response.content)
                    except Exception as e:
                        # Log or handle download error if needed
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=mega_booklet_correction_images.zip'
        return response

    download_selected_images.short_description = "Download selected images as zip"

    def serial_number(self, obj):
        try:
            queryset = self.model.objects.order_by('id')
            index = list(queryset).index(obj)
            return index + 1
        except Exception:
            return '-'

    serial_number.short_description = 'S.No'

    @admin.action(description='Export selected mega booklet correction requests to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mega Booklet Correction Requests"

        headers = ['S.No', 'Candidate Name', 'DOB', 'City', 'WhatsApp Number', 'Booklet Serial', 'Booklet Name', 'Correction Description', 'Submission Time']
        ws.append(headers)

        # Set column widths
        column_widths = [8, 25, 15, 20, 20, 15, 25, 50, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ordered_queryset = queryset.order_by('id')

        for idx, obj in enumerate(ordered_queryset, start=1):
            row = [
                idx,
                obj.candidate_name,
                obj.dob,
                obj.city,
                obj.whatsapp_number,
                obj.booklet_serial,
                obj.booklet_name,
                obj.correction_description,
                obj.submission_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(obj.submission_time, 'strftime') else obj.submission_time or '',
            ]
            ws.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=mega_booklet_correction_requests.xlsx'
        wb.save(response)
        return response

    actions = ['export_selected_to_excel', 'download_selected_images']
