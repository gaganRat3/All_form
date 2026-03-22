"""
Utility functions for exporting Get-Together registrations
"""
import csv
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


def export_get_together_to_csv(registrations):
    """Export registrations to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=get_together_registrations.csv'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Candidate Name', 'Gender', 'Date of Birth', 'City', 'WhatsApp No.', 'Members', 'Submitted At'])
    
    for registration in registrations:
        writer.writerow([
            registration.id,
            registration.candidate_name,
            registration.get_gender_display(),
            registration.dob.strftime('%Y-%m-%d'),
            registration.city,
            registration.whatsapp,
            registration.get_members_display(),
            registration.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])
    
    return response


def export_get_together_to_excel(registrations):
    """Export registrations to Excel format with styling"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Get-Together Registrations"
    
    # Define headers
    headers = ['ID', 'Candidate Name', 'Gender', 'Date of Birth', 'City', 'WhatsApp No.', 'Members', 'Submitted At']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='A8652C', end_color='A8652C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    column_widths = [8, 20, 12, 15, 15, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add data rows
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num, registration in enumerate(registrations, 2):
        row_data = [
            registration.id,
            registration.candidate_name,
            registration.get_gender_display(),
            registration.dob.strftime('%Y-%m-%d'),
            registration.city,
            registration.whatsapp,
            registration.get_members_display(),
            registration.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
        ]
        ws.append(row_data)
        
        # Apply styling to data rows
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Center align certain columns
            if col_num in [1, 3, 7]:  # ID, Gender, Members
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add summary statistics
    summary_row = len(registrations) + 3
    ws[f'A{summary_row}'] = 'Summary'
    ws[f'A{summary_row}'].font = Font(bold=True, size=11)
    
    ws[f'A{summary_row + 1}'] = f'Total Registrations: {len(registrations)}'
    
    # Count by gender
    try:
        from django.db.models import Count
        gender_counts = {}
        for reg in registrations:
            gender = reg.get_gender_display()
            gender_counts[gender] = gender_counts.get(gender, 0) + 1
        
        row = summary_row + 2
        for gender, count in gender_counts.items():
            ws[f'A{row}'] = f'{gender}: {count}'
            row += 1
    except:
        pass
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=get_together_registrations.xlsx'
    wb.save(response)
    return response
