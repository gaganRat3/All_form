from django.contrib import admin
from .models import AdvancePassBooking, BookletLibrarySubmission, FortyPlusSammelan, BhudevKalakaar2026Registration, SaurasthraKutchSammelan, CandidateBiodata
# Register CandidateBiodata in admin
@admin.register(CandidateBiodata)
class CandidateBiodataAdmin(admin.ModelAdmin):
    list_display = ('candidate_name', 'gender', 'registrant_mobile', 'candidate_current_city', 'dob', 'marital_status', 'birth_time', 'birth_place', 'residence_area_category')
    search_fields = ('candidate_name', 'registrant_mobile', 'candidate_current_city', 'birth_place')
    list_filter = ('gender', 'marital_status', 'residence_area_category')
from .admin_37th_sammelan import *

@admin.register(FortyPlusSammelan)
class FortyPlusSammelanAdmin(admin.ModelAdmin):
    list_display = [
        'serial_number', 'name', 'gender', 'dob', 'marital', 'disability', 'tob', 'birthPlace', 'city', 'country', 'visa', 'height', 'weight',
        'education', 'educationDetail', 'occupationCat', 'occupationDetails', 'salary', 'shani', 'hobbies', 'father', 'mother',
        'fatherWp', 'motherWp', 'caste', 'gotra', 'kuldevi', 'siblings', 'eating_habbits', 'alcohol', 'smoke', 'other_habbit',
        'legal_case', 'locChoice', 'ageGap', 'eduChoice', 'otherChoice', 'who', 'regMobile', 'resCat', 'nadi', 'email', 'whatsapp',
        'photo', 'declaration', 'submitted_at'
    ]
    search_fields = ['name', 'email', 'regMobile', 'city']
    list_filter = ['gender', 'marital', 'city']
    actions = ['export_excel_with_images', 'export_excel_without_images', 'download_images_zip']

    def serial_number(self, obj):
        """Display serial number based on submitted_at timestamp"""
        all_ids = list(FortyPlusSammelan.objects.all().order_by('submitted_at').values_list('id', flat=True))
        try:
            position = all_ids.index(obj.id)
            return position + 1
        except ValueError:
            return 0

        sorted_queryset = list(queryset.order_by('-submitted_at'))
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in sorted_queryset:
                if obj.photo:
                    try:
                        img_path = obj.photo.path
                        if os.path.exists(img_path):
                            serial_number = id_to_serial.get(obj.id, 0)
                            filename = f"{serial_number}_{obj.name.replace(' ', '_')}{os.path.splitext(img_path)[1]}"
                            zip_file.write(img_path, filename)
                    except Exception as e:
                        print(f"Error adding image to zip: {e}")
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=physical_form_images.zip'
        return response


@admin.register(SaurasthraKutchSammelan)
class SaurasthraKutchSammelanAdmin(admin.ModelAdmin):
    list_display = [
        'serial_number', 'name', 'gender', 'dob', 'marital', 'disability', 'tob', 'birthPlace', 'city', 'country', 'visa', 'height', 'weight',
        'education', 'educationDetail', 'occupationCat', 'occupationDetails', 'salary', 'shani', 'hobbies', 'father', 'mother',
        'fatherWp', 'motherWp', 'caste', 'gotra', 'kuldevi', 'siblings', 'eating_habbits', 'alcohol', 'smoke', 'other_habbit',
        'legal_case', 'locChoice', 'ageGap', 'eduChoice', 'otherChoice', 'who', 'regMobile', 'resCat', 'nadi', 'email', 'whatsapp',
        'photo', 'declaration', 'submitted_at'
    ]
    search_fields = ['name', 'email', 'regMobile', 'city']
    list_filter = ['gender', 'marital', 'city', 'resCat']
    actions = ['export_excel_with_images', 'export_excel_without_images', 'download_images_zip']

    def serial_number(self, obj):
        """Display serial number based on submitted_at timestamp"""
        all_ids = list(SaurasthraKutchSammelan.objects.all().order_by('submitted_at').values_list('id', flat=True))
        try:
            position = all_ids.index(obj.id)
            return position + 1
        except ValueError:
            return 0

# Register BookletLibrarySubmission in admin
@admin.register(BookletLibrarySubmission)
class BookletLibrarySubmissionAdmin(admin.ModelAdmin):
    list_display = ('name', 'city', 'whatsapp', 'email', 'gender', 'dob', 'submitted_at')
    search_fields = ('name', 'city', 'whatsapp', 'email')
    list_filter = ('city', 'gender', 'submitted_at')
from .models import CourierBooklet35thBooking
from .models_picnic import PicnicRegistration
from django.utils.html import format_html
from .models_booklet_camp_adv import BookletCampAdvBooking

# Register CourierBooklet35thBooking in admin
@admin.register(CourierBooklet35thBooking)
class CourierBooklet35thBookingAdmin(admin.ModelAdmin):
    list_display = ('name', 'city', 'whatsapp_number', 'email', 'girls_booklet_with', 'boys_booklet_with', 'total_amount', 'payment_screenshot_preview', 'created_at')
    list_filter = ['created_at', 'city', 'girls_booklet_with', 'boys_booklet_with']
    search_fields = ['name', 'city', 'whatsapp_number', 'email']
    readonly_fields = ['payment_screenshot_preview', 'created_at']
    actions = ['export_selected_to_excel']

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html(f'<img src="{obj.payment_screenshot.url}" style="max-height: 100px; max-width: 100px;" />')
        return "-"
    payment_screenshot_preview.short_description = 'Payment Screenshot'

    @admin.action(description='Export selected bookings to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courier Booklet Bookings"

        headers = ['Name', 'City', 'WhatsApp Number', 'Email', 'Girls Booklet', 'Boys Booklet', 'Courier Address', 'Total Amount', 'Payment Screenshot', 'Created At']
        ws.append(headers)

        column_widths = [20, 20, 20, 30, 15, 15, 40, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.city,
                obj.whatsapp_number,
                obj.email,
                'Yes' if obj.girls_booklet_with else 'No',
                'Yes' if obj.boys_booklet_with else 'No',
                obj.courier_address,
                obj.total_amount,
                '',  # Placeholder for image
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_path = obj.payment_screenshot.path
                    pil_img = PILImage.open(img_path)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(9)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception as e:
                    pass
            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=courier_booklet_bookings.xlsx'
        wb.save(response)
        return response

# Register AdvancePassBooking in admin
@admin.register(AdvancePassBooking)
class AdvancePassBookingAdmin(admin.ModelAdmin):
    list_display = [
        'name', 'city', 'whatsapp_number', 'email',
        'entry_token_quantity', 'unlimited_buffet_quantity',
        'total_amount', 'payment_screenshot_preview', 'created_at'
    ]
    list_filter = ['created_at']
    search_fields = ['name', 'city', 'whatsapp_number', 'email']
    readonly_fields = ['payment_screenshot', 'payment_screenshot_preview']

    actions = ['export_selected_to_excel']

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html(f'<img src="{obj.payment_screenshot.url}" style="max-height: 100px; max-width: 100px;" />')
        return "-"
    payment_screenshot_preview.short_description = 'Payment Screenshot'

    @admin.action(description='Export selected advance pass bookings to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        import requests
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advance Pass Bookings"

        # Define headers
        headers = ['Name', 'City', 'WhatsApp Number', 'Email', 'Entry Token Qty', 'Buffet Qty', 'Total Amount', 'Payment Screenshot', 'Created At']
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 30, 15, 15, 15, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.city,
                obj.whatsapp_number,
                obj.email,
                obj.entry_token_quantity,
                obj.unlimited_buffet_quantity,
                obj.total_amount,
                '',  # Placeholder for image
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_path = obj.payment_screenshot.path
                    pil_img = PILImage.open(img_path)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(9)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception as e:
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advance_pass_bookings.xlsx'
        wb.save(response)
        return response
from django.utils.html import format_html
# Import mark_safe at the top so it is available for all uses
from django.utils.safestring import mark_safe
# linebreaksbr is a template filter available in django.template.defaultfilters
from django.template.defaultfilters import linebreaksbr
from django.http import HttpResponseRedirect
from .models_business_directory import BusinessDirectoryEntry
from .models import SpotAdvanceBookletBooking, AdvanceBookletBooking, BncBnfApplication, GarbaPassRegistration, ParticipantRegistration, AudienceRegistration
from .models_booklet_camp_adv import BookletCampAdvBooking

# Register PicnicRegistration in admin
@admin.register(PicnicRegistration)
class PicnicRegistrationAdmin(admin.ModelAdmin):
    list_display = ['filler_name', 'contact_number', 'email_address', 'city', 'candidate_name', 'gender', 'relation', 'formatted_dob', 'persons', 'is_coming', 'submitted_at']
    list_filter = ['submitted_at', 'is_coming', 'gender']
    search_fields = ['filler_name', 'contact_number', 'email_address', 'candidate_name', 'city']
    readonly_fields = ['submitted_at', 'formatted_dob']
    
    fieldsets = (
        ('Filler Information', {
            'fields': ('filler_name', 'contact_number', 'email_address', 'city')
        }),
        ('Candidate Information', {
            'fields': ('candidate_name', 'gender', 'relation', 'dob')
        }),
        ('Picnic Details', {
            'fields': ('persons', 'is_coming')
        }),
        ('Meta', {
            'fields': ('submitted_at',),
            'classes': ('collapse',)
        }),
    )
    
    def formatted_dob(self, obj):
        return obj.dob.strftime('%d-%m-%Y') if obj.dob else '-'
    formatted_dob.short_description = 'Date of Birth'
    
    actions = ['export_selected_to_excel']
    
    @admin.action(description='Export selected picnic registrations to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Picnic Registrations"
        
        # Define headers
        headers = ['Filler Name', 'Contact Number', 'City', 'Candidate Name', 'Gender', 'Relation', 'Date of Birth', 'Number of Persons', 'Is Coming', 'Submitted At']
        ws.append(headers)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Set column widths
        column_widths = [20, 20, 20, 20, 12, 15, 15, 15, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add data rows
        row_num = 2
        for obj in queryset:
            row = [
                obj.filler_name,
                obj.contact_number,
                obj.city,
                obj.candidate_name,
                obj.gender,
                obj.relation,
                obj.dob.strftime('%d-%m-%Y') if obj.dob else '',
                obj.persons,
                obj.is_coming,
                str(obj.submitted_at),
            ]
            ws.append(row)
            row_num += 1
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=picnic_registrations.xlsx'
        wb.save(response)
        return response

# Register AdvanceBookletBooking in admin panel
@admin.register(AdvanceBookletBooking)
class AdvanceBookletBookingAdmin(admin.ModelAdmin):
    list_display = ['name', 'city', 'whatsapp_number', 'email', 'girls_booklet_with', 'boys_booklet_with', 'total_amount', 'courier_address', 'payment_screenshot', 'created_at']
    search_fields = ['name', 'city', 'whatsapp_number', 'email']
    list_filter = ['created_at']
    readonly_fields = ['created_at']
from .models import SammelanPaymentForm
from .models import BirthdayFormSubmission, StorySubmission
from .models import DivorceSammelanForm

from django.utils.translation import gettext_lazy as _
from django.contrib.admin import SimpleListFilter
from openpyxl import Workbook
from django.shortcuts import get_object_or_404






















@admin.register(StorySubmission)
class StorySubmissionAdmin(admin.ModelAdmin):
    # Show all relevant fields for the happy story form in the list view so
    # admins can quickly see requested details.
    # Display order updated to follow requested sequence:
    # Boy Candidate Name, Boy Birth Date, Boy City, Girl Candidate Name, Girl Birth Date, Girl City, Who is filling, Mobile Number, WhatsApp Number, Relationship Status, Function Date, Message, Photos, Submitted At
    list_display = [
        'boy_name',
        'boy_birth_date',
        'boy_city',
        'girl_name',
        'girl_birth_date',
        'girl_city',
        'who_is_filling',
        'mobile_number',
        'whatsapp_number',
        'relationship_status',
        'function_date',
        'full_story',
        'image_preview',
        'submitted_at',
    ]
    readonly_fields = ['image_preview', 'submitted_at']
    search_fields = ['boy_name', 'girl_name', 'mobile_number', 'whatsapp_number', 'who_is_filling']
    list_filter = ['relationship_status', 'submitted_at']
    actions = ['export_as_excel', 'download_images_zip']
    # Make admin form fields explicit so the change form shows expected inputs (labels follow model verbose_name)
    fields = (
        'boy_name', 'boy_birth_date', 'boy_city',
        'girl_name', 'girl_birth_date', 'girl_city',
        'who_is_filling', 'mobile_number', 'whatsapp_number', 'relationship_status',
        'function_date', 'message', 'image', 'image_preview', 'submitted_at'
    )
    list_display_links = ('girl_name', 'boy_name')
    ordering = ('-submitted_at',)
    
    def full_story(self, obj):
        """Return the full story text (from `message` or legacy `details`) with line breaks preserved."""
        text = obj.message or ''
        if not text:
            return '-'
        return format_html('<div style="white-space:normal; max-width:900px;">{}</div>', linebreaksbr(text))
    full_story.short_description = 'Share Your Wedding Story'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<a href="{}" target="_blank">{}</a>', obj.image.url, obj.image.name.split('/')[-1])
        return '-'
    image_preview.short_description = 'Image'

    @admin.action(description='Export selected as Excel')
    def export_as_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Story Submissions'

        # Include all fields shown in admin list_display and form
        headers = [
            'Boy Name', 'Boy Birth Date', 'Boy City',
            'Girl Name', 'Girl Birth Date', 'Girl City',
            'Who Is Filling', 'Mobile Number', 'WhatsApp Number',
            'Relationship Status', 'Function Date', 'Message', 'Image', 'Submitted At'
        ]
        ws.append(headers)

        for obj in queryset:
            row = [
                obj.boy_name or '',
                obj.boy_birth_date.strftime('%Y-%m-%d') if getattr(obj, 'boy_birth_date', None) else '',
                obj.boy_city or '',
                obj.girl_name or '',
                obj.girl_birth_date.strftime('%Y-%m-%d') if getattr(obj, 'girl_birth_date', None) else '',
                obj.girl_city or '',
                obj.who_is_filling or '',
                obj.mobile_number or '',
                obj.whatsapp_number or '',
                obj.relationship_status or '',
                obj.function_date.isoformat() if getattr(obj, 'function_date', None) else '',
                (obj.message or '')[:32767],  # Excel cell limit
                (obj.image.url if getattr(obj, 'image', None) else ''),
                obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(obj, 'submitted_at', None) else ''
            ]
            ws.append(row)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=story_submissions.xlsx'
        wb.save(response)
        return response
    export_as_excel.short_description = 'Export selected as Excel'

    @admin.action(description='Download images as ZIP')
    def download_images_zip(self, request, queryset):
        import io, zipfile, os
        from django.http import HttpResponse
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for obj in queryset:
                if obj.image:
                    try:
                        img_path = obj.image.path
                        if os.path.exists(img_path):
                            zip_file.write(img_path, f"story_images/{os.path.basename(img_path)}")
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=story_images.zip'
        return response
    download_images_zip.short_description = 'Download images as ZIP'

    


# As an extra safety, ensure the model is registered even if decorators changed or reloads
try:
    from .models import StorySubmission as _StorySubmission
    if _StorySubmission not in admin.site._registry:
        admin.site.register(_StorySubmission, StorySubmissionAdmin)
except Exception:
    # If registration fails (already registered or import error), ignore silently
    pass


@admin.register(GarbaPassRegistration)
class GarbaPassRegistrationAdmin(admin.ModelAdmin):
    actions = ['export_selected_to_excel', 'export_payment_screenshots_zip']

    def passes_display(self, obj):
        if obj.passes:
            passes_list = []
            for p in obj.passes:
                passes_list.append(f"{p.get('type', '')} x{p.get('quantity', 0)} (₹{p.get('amount', 0)})")
            return ", ".join(passes_list)
        return "No passes"
    passes_display.short_description = "Category of Passes"

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html('<img src="{}" style="max-height:80px; max-width:120px; border-radius:8px;" />', obj.payment_screenshot.url)
        return "No screenshot"

    payment_screenshot_preview.short_description = "Upload Payment Screenshot"

    @admin.action(description='Export selected garba pass registrations to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import requests
        from PIL import Image as PILImage
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Garba Pass Registrations"

        # Define headers
        headers = [
            'Full Name', 'Date of Birth', 'Residence City', 'WhatsApp Number',
            'Passes', 'Subtotal (₹)', 'Created At', 'Payment Screenshot'
        ]
        ws.append(headers)

        # Set column widths
        column_widths = [25, 20, 25, 20, 30, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            # Format passes information
            passes_info = ""
            if obj.passes:
                passes_list = []
                for p in obj.passes:
                    passes_list.append(f"{p.get('type', '')} x{p.get('quantity', 0)} (₹{p.get('amount', 0)})")
                passes_info = ", ".join(passes_list)

            row = [
                obj.full_name,
                obj.date_of_birth,
                obj.residence_city,
                obj.whatsapp_number,
                passes_info,
                obj.subtotal,
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                '',  # Placeholder for image
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_url = obj.payment_screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(8)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=garba_pass_registrations.xlsx'
        wb.save(response)
        return response

    @admin.action(description='Export payment screenshots of selected garba registrations as ZIP')
    def export_payment_screenshots_zip(self, request, queryset):
        import zipfile
        import os
        from io import BytesIO
        import re

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in queryset:
                if obj.payment_screenshot:
                    try:
                        img_path = obj.payment_screenshot.path
                        ext = os.path.splitext(img_path)[1].lower()
                        if os.path.exists(img_path):
                            # Sanitize name and whatsapp number for filename
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.full_name.strip())
                            safe_whatsapp = re.sub(r'[^0-9]', '', obj.whatsapp_number)
                            filename = f"{safe_name}_{safe_whatsapp}{ext}"
                            with open(img_path, 'rb') as img_file:
                                img_data = img_file.read()
                            zip_file.writestr(filename, img_data)
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=garba_pass_payment_screenshots.zip'
        return response

    list_display = ['full_name', 'date_of_birth', 'residence_city', 'whatsapp_number', 'passes_display', 'subtotal', 'payment_screenshot_preview', 'created_at']
    search_fields = ['full_name', 'whatsapp_number', 'residence_city']


@admin.register(BirthdayFormSubmission)
class BirthdayFormSubmissionAdmin(admin.ModelAdmin):
    list_display = ['candidate_name', 'birth_date', 'mobile_1', 'mobile_2', 'mobile_3', 'city', 'created_at']
    search_fields = ['candidate_name', 'mobile_1', 'mobile_2', 'mobile_3', 'city']
    readonly_fields = ['created_at']
    actions = ['export_selected_birthday_to_excel']

    @admin.action(description='Export selected birthday submissions to Excel')
    def export_selected_birthday_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Birthday Submissions'

        headers = ['ID', 'Candidate Name', 'Birth Date', 'Mobile 1', 'Mobile 2', 'Mobile 3', 'City', 'Submitted At']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, obj in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=obj.id)
            ws.cell(row=row_num, column=2, value=obj.candidate_name)
            ws.cell(row=row_num, column=3, value=obj.birth_date.strftime('%Y-%m-%d') if obj.birth_date else '')
            ws.cell(row=row_num, column=4, value=obj.mobile_1)
            ws.cell(row=row_num, column=5, value=obj.mobile_2 or '')
            ws.cell(row=row_num, column=6, value=obj.mobile_3 or '')
            ws.cell(row=row_num, column=7, value=obj.city or '')
            ws.cell(row=row_num, column=8, value=obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '')

        # Auto-size columns (simple heuristic)
        for i, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=birthday_submissions.xlsx'
        wb.save(response)
        return response

@admin.register(AudienceRegistration)
class AudienceRegistrationAdmin(admin.ModelAdmin):
    actions = ['export_selected_to_excel', 'export_payment_screenshots_zip']

    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html('<img src="{}" style="max-height:80px; max-width:120px; border-radius:8px;" />', obj.payment_screenshot.url)
        return "No screenshot"

    payment_screenshot_preview.short_description = "Payment Screenshot"

    def full_name_display(self, obj):
        return f"{obj.first_name} {obj.last_name}"
    
    full_name_display.short_description = "Full Name"

    @admin.action(description='Export selected audience registrations to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import requests
        from PIL import Image as PILImage
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audience Registrations"

        # Define headers
        headers = [
            'First Name', 'Last Name', 'Email', 'Phone Number',
            'Ticket Quantity', 'Total Amount (₹)', 'Created At', 'Payment Screenshot'
        ]
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 30, 20, 15, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.first_name,
                obj.last_name,
                obj.email,
                obj.phone_number,
                obj.ticket_quantity,
                float(obj.total_amount),
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                '',  # Placeholder for image
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if obj.payment_screenshot:
                try:
                    img_url = obj.payment_screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(8)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=audience_registrations.xlsx'

        # Save workbook to response
        wb.save(response)
        return response

    @admin.action(description='Export payment screenshots of selected audience registrations as ZIP')
    def export_payment_screenshots_zip(self, request, queryset):
        import zipfile
        import os
        from io import BytesIO
        import re

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in queryset:
                if obj.payment_screenshot:
                    try:
                        img_path = obj.payment_screenshot.path
                        ext = os.path.splitext(img_path)[1].lower()
                        if os.path.exists(img_path):
                            # Sanitize name and phone number for filename
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', f"{obj.first_name}_{obj.last_name}".strip())
                            safe_phone = re.sub(r'[^0-9]', '', obj.phone_number)
                            filename = f"{safe_name}_{safe_phone}{ext}"
                            with open(img_path, 'rb') as img_file:
                                img_data = img_file.read()
                            zip_file.writestr(filename, img_data)
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=audience_payment_screenshots.zip'
        return response

    list_display = ['full_name_display', 'email', 'phone_number', 'ticket_quantity', 'total_amount', 'payment_screenshot_preview', 'created_at']
    search_fields = ['first_name', 'last_name', 'email', 'phone_number']
    list_filter = ['ticket_quantity', 'created_at']
    readonly_fields = ['created_at']
    ordering = ['-created_at']

# Register BNCF Application model for admin panel
@admin.register(BncBnfApplication)
class BncBnfApplicationAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'gender', 'date_of_birth', 'marriage_status', 'whatsapp_number', 'phone_number', 'education', 'occupation', 'current_city', 'area_name', 'home_address']
    search_fields = ['full_name', 'whatsapp_number', 'phone_number', 'current_city', 'area_name']
from .models_astrology import AstrologyFormSubmission

@admin.register(SpotAdvanceBookletBooking)
class SpotAdvanceBookletBookingAdmin(admin.ModelAdmin):
    actions = ['export_selected_to_excel', 'export_payment_screenshots_zip']

    @admin.action(description='Export selected spot advance booklet bookings to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import requests
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Spot Advance Booklet Bookings"

        # Define headers
        headers = [
            'Name', 'City', 'Whatsapp Number', 'Email',
            'Saurashtra Booklet', 'Gujarat Girls Booklet', 'Gujarat Boys Booklet',
            'NRI Booklet', 'Mumbai Booklet', 'Divorce/Widow Booklet',
            'Booklet Camp City', 'Candidate Name & DOB',
            'Total Amount', 'Created At', 'Payment Screenshot'
        ]
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 30, 20, 20, 20, 20, 20, 20, 20, 30, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.city,
                obj.whatsapp_number,
                obj.email,
                obj.saurashtra_booklet,
                obj.gujarat_girls_booklet,
                obj.gujarat_boys_booklet,
                obj.nri_booklet,
                obj.mumbai_booklet,
                obj.divorce_widow_booklet,
                obj.booklet_camp_city,
                obj.candidate_name_dob,
                obj.total_amount,
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                '',  # Placeholder for image
            ]
            ws.append(row)

            # Embed payment screenshot image if available
            if hasattr(obj, 'payment_screenshot') and obj.payment_screenshot:
                try:
                    img_url = obj.payment_screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(15)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=spot_advance_booklet_bookings.xlsx'

        # Save workbook to response
        wb.save(response)
        return response

    @admin.action(description='Export payment screenshots of selected spot advance booklet bookings as ZIP')
    def export_payment_screenshots_zip(self, request, queryset):
        import zipfile
        import os
        from io import BytesIO
        import re

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for obj in queryset:
                if hasattr(obj, 'payment_screenshot') and obj.payment_screenshot:
                    try:
                        img_path = obj.payment_screenshot.path
                        ext = os.path.splitext(img_path)[1].lower()
                        if os.path.exists(img_path):
                            # Sanitize name and whatsapp number for filename
                            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', obj.name.strip())
                            safe_whatsapp = re.sub(r'[^0-9]', '', obj.whatsapp_number)
                            filename = f"{safe_name}_{safe_whatsapp}{ext}"
                            with open(img_path, 'rb') as img_file:
                                img_data = img_file.read()
                            zip_file.writestr(filename, img_data)
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=spot_advance_booklet_payment_screenshots.zip'
        return response
    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html('<img src="{}" style="max-height:80px; max-width:120px; border-radius:8px;" />', obj.payment_screenshot.url)
        return "No screenshot uploaded"

    payment_screenshot_preview.short_description = "Payment Screenshot"

    list_display = [
        'name', 'city', 'whatsapp_number', 'email',
        'saurashtra_booklet', 'gujarat_girls_booklet', 'gujarat_boys_booklet',
        'nri_booklet', 'mumbai_booklet', 'divorce_widow_booklet',
        'booklet_camp_city', 'candidate_name_dob',
        'total_amount', 'created_at', 'payment_screenshot_preview'
    ]
    list_filter = [
        'saurashtra_booklet', 'gujarat_girls_booklet', 'gujarat_boys_booklet',
        'nri_booklet', 'mumbai_booklet', 'divorce_widow_booklet',
        'booklet_camp_city', 'created_at'
    ]
    search_fields = ['name', 'city', 'whatsapp_number', 'email', 'booklet_camp_city', 'candidate_name_dob']
from .models import ZoomRegistration
import io
import zipfile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
@admin.register(BusinessDirectoryEntry)
class BusinessDirectoryEntryAdmin(admin.ModelAdmin):
    def ownerPhoto_preview(self, obj):
        if hasattr(obj, 'ownerPhoto') and obj.ownerPhoto:
            return format_html(
                '<img src="{}" style="max-height:80px; max-width:120px; border-radius:8px;" /><br>'
                '<a href="{}" download style="color: #007bff; text-decoration: underline;">Download Owner Photo</a>',
                obj.ownerPhoto.url, obj.ownerPhoto.url
            )
        return "No owner photo uploaded"

    list_display = [
        'businessName', 'ownerName', 'services', 'service_locations', 'address',
        'phone', 'whatsapp', 'email', 'website',
        'logo_preview', 'ownerPhoto_preview', 'created_at', 'updated_at'
    ]
    search_fields = ('businessName', 'ownerName', 'phone', 'email')
    readonly_fields = ('logo_preview', 'ownerPhoto_preview')
    fields = (
        'businessName', 'ownerName', 'services', 'service_locations', 'address',
        'phone', 'whatsapp', 'email', 'website',
        'logo', 'logo_preview', 'ownerPhoto', 'ownerPhoto_preview'
    )

    actions = ['export_as_excel', 'download_images_zip']

    def logo_preview(self, obj):
        if obj.logo:
            return format_html(
                '<img src="{}" style="max-height:80px; max-width:120px; border-radius:8px;" /><br>'
                '<a href="{}" download style="color: #007bff; text-decoration: underline;">Download Logo</a>',
                obj.logo.url, obj.logo.url
            )
        return "No logo uploaded"

    def export_as_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Business Directory"
        headers = [
            'Business Name', 'Owner Name', 'Services', 'Service Locations', 'Address',
            'Phone', 'Whatsapp', 'Email', 'Website', 'Logo', 'Owner Photo', 'Created At', 'Updated At'
        ]
        ws.append(headers)

        # Set column widths for better image display
        col_widths = [20, 20, 30, 25, 30, 15, 15, 25, 25, 15, 15, 20, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            ws.append([
                obj.businessName, obj.ownerName, obj.services, obj.service_locations, obj.address,
                obj.phone, obj.whatsapp, obj.email, obj.website,
                '',  # Placeholder for logo image
                '',  # Placeholder for owner photo image
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                obj.updated_at.strftime('%Y-%m-%d %H:%M:%S') if obj.updated_at else ''
            ])

            # Embed logo image if available
            if obj.logo:
                try:
                    img_path = obj.logo.path
                    pil_img = PILImage.open(img_path)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(10)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass

            # Embed owner photo image if available
            if obj.ownerPhoto:
                try:
                    img_path = obj.ownerPhoto.path
                    pil_img = PILImage.open(img_path)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(11)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=business_directory.xlsx'
        wb.save(response)
        return response
    export_as_excel.short_description = "Export selected as Excel"

    def download_images_zip(self, request, queryset):
        import io
        import zipfile
        from django.http import HttpResponse
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for obj in queryset:
                if obj.logo:
                    logo_path = obj.logo.path
                    zip_file.write(logo_path, f"logos/{obj.logo.name.split('/')[-1]}")
                if obj.ownerPhoto:
                    owner_path = obj.ownerPhoto.path
                    zip_file.write(owner_path, f"owner_photos/{obj.ownerPhoto.name.split('/')[-1]}")
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type="application/zip")
        response['Content-Disposition'] = 'attachment; filename=business_directory_images.zip'
        return response
    download_images_zip.short_description = "Download images as ZIP"




@admin.action(description='Export selected astrology form submissions to Excel')
def export_selected_astrology_to_excel(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Astrology Form Submissions"

    headers = ['Name', 'Whatsapp', 'City', 'Candidate Name', 'Candidate DOB', 'Candidate Birth Time', 'Candidate Birth Place', 'Description', 'Submitted At']
    ws.append(headers)

    column_widths = [20, 20, 20, 25, 20, 20, 25, 50, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for obj in queryset:
        row = [
            obj.name,
            obj.whatsapp,
            obj.city,
            obj.candidateName,
            obj.candidateDOB,
            obj.candidateBirthTime,

            obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else '',
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=astrology_form_submissions.xlsx'
    wb.save(response)
    return response

@admin.register(AstrologyFormSubmission)
class AstrologyFormSubmissionAdmin(admin.ModelAdmin):
    list_display = ['name', 'whatsapp', 'city', 'candidateName', 'candidateDOB', 'candidateBirthTime', 'candidateBirthPlace', 'description', 'submitted_at']
    search_fields = ['name', 'whatsapp', 'city', 'candidateName', 'candidateDOB', 'candidateBirthTime', 'candidateBirthPlace']
    readonly_fields = ['submitted_at']
    actions = [export_selected_astrology_to_excel]

@admin.register(ZoomRegistration)
class ZoomRegistrationAdmin(admin.ModelAdmin):
    list_display = ['name', 'whatsapp', 'city', 'submitted_at', 'download_link', 'screenshot_preview']
    readonly_fields = ['download_link', 'screenshot_preview']  # show on detail page
    search_fields = ['name', 'whatsapp', 'city']  # Added search fields for admin search functionality
    actions = ['export_as_excel', 'download_screenshots_zip']

    def download_screenshots_zip(self, request, queryset):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for obj in queryset:
                if obj.screenshot:
                    file_path = obj.screenshot.path
                    file_name = obj.screenshot.name.split('/')[-1]
                    with open(file_path, 'rb') as f:
                        zip_file.writestr(file_name, f.read())
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=payment_screenshots.zip'
        return response

    download_screenshots_zip.short_description = "Download Selected Screenshots as ZIP"

    def download_link(self, obj):
        if obj.screenshot:
            return format_html(
                '<a href="{}" download target="_blank">Download Screenshot</a>',
                obj.screenshot.url
            )
        return "-"
    download_link.short_description = "Payment Screenshot"

    def screenshot_preview(self, obj):
        if obj.screenshot:
            return format_html('<img src="{}" width="100"/>', obj.screenshot.url)
        return "-"
    screenshot_preview.short_description = "Preview"

    def export_as_excel(self, request, queryset):
        import requests
        from PIL import Image as PILImage
        from io import BytesIO
        import logging
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zoom Registrations"

        # Define the headers
        headers = ['Name', 'WhatsApp', 'City', 'Screenshot', 'Submitted At']
        ws.append(headers)

        row_num = 2
        for obj in queryset:
            ws.append([
                obj.name,
                obj.whatsapp,
                obj.city,
                '',  # Placeholder for image
                obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])

            # Embed screenshot image if available
            if obj.screenshot:
                try:
                    img_url = obj.screenshot.url
                    if img_url.startswith('/'):
                        base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                        img_url = base_url + img_url
                    response_img = requests.get(img_url)
                    response_img.raise_for_status()
                    img_data = BytesIO(response_img.content)
                    pil_img = PILImage.open(img_data)
                    img_byte_arr = BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    img = OpenpyxlImage(img_byte_arr)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"{get_column_letter(4)}{row_num}"
                    ws.add_image(img)
                    ws.row_dimensions[row_num].height = 60
                except Exception as e:
                    logging.error(f"Failed to embed image in Excel export for ZoomRegistration id {obj.id}: {e}")
            row_num += 1

        # Set column widths (optional)
        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=zoom_registrations.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "Export Selected to Excel"


from django.contrib import admin
from .models_technical_support import TechnicalSupportRequest, SupportAttachment

class TechnicalSupportRequestAdmin(admin.ModelAdmin):
    actions = ['export_selected_to_excel']

    list_display = ('first_name', 'last_name', 'city', 'phone', 'whatsapp_number', 'category', 'description', 'image_preview', 'created_at')
    fields = ('first_name', 'last_name', 'city', 'phone', 'whatsapp_number', 'category', 'description', 'created_at', 'image_preview')
    readonly_fields = ('created_at', 'image_preview')

    @admin.action(description='Export selected technical support requests to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        import requests
        from PIL import Image as PILImage
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Technical Support Requests"

        # Define headers
        headers = [
            'First Name', 'Last Name', 'City', 'Phone', 'WhatsApp Number',
            'Category', 'Description', 'Created At', 'Attachments'
        ]
        ws.append(headers)

        # Set column widths
        column_widths = [20, 20, 20, 20, 20, 20, 50, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.first_name,
                obj.last_name,
                obj.city,
                obj.phone,
                obj.whatsapp_number,
                obj.get_category_display(),  # Display the choice label
                obj.description,
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
                '',  # Placeholder for image
            ]
            ws.append(row)

            # Embed first image attachment if available
            attachments = obj.attachments.all()
            image_embedded = False
            for attachment in attachments:
                if attachment.file.url.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                    try:
                        img_url = attachment.file.url
                        if img_url.startswith('/'):
                            base_url = 'https://bhudevnetwork.pythonanywhere.com/'  # Replace with your actual domain or base URL
                            img_url = base_url + img_url
                        response_img = requests.get(img_url)
                        response_img.raise_for_status()
                        img_data = BytesIO(response_img.content)
                        pil_img = PILImage.open(img_data)
                        img_byte_arr = BytesIO()
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        img = OpenpyxlImage(img_byte_arr)
                        img.width = 80
                        img.height = 80
                        img.anchor = f"{get_column_letter(9)}{row_num}"
                        ws.add_image(img)
                        ws.row_dimensions[row_num].height = 60
                        image_embedded = True
                        break  # Embed only the first image
                    except Exception:
                        pass
            if not image_embedded:
                # If no image, put text
                ws.cell(row=row_num, column=9).value = 'No images' if not attachments else 'Non-image attachments'

            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=technical_support_requests.xlsx'

        # Save workbook to response
        wb.save(response)
        return response

    def image_preview(self, obj):
        attachments = obj.attachments.all()
        if attachments:
            previews = []
            for a in attachments:
                if a.file.url.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                    previews.append(f'<img src="{a.file.url}" style="max-width:80px;max-height:80px;border-radius:8px;margin-right:8px;" />')
                else:
                    previews.append(f'<a href="{a.file.url}" target="_blank">{a.filename}</a>')
            return format_html(''.join(previews))
        return "No images"
    image_preview.short_description = "Images"

admin.site.register(TechnicalSupportRequest, TechnicalSupportRequestAdmin)
admin.site.register(SupportAttachment)
from .admin_karmkand_directory import *


@admin.register(ParticipantRegistration)
class ParticipantRegistrationAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'email', 'phone_number', 'events_display', 'subtotal', 'payment_screenshot_thumb', 'submission_time']
    list_filter = ['submission_time', 'subtotal']
    search_fields = ['first_name', 'last_name', 'email', 'phone_number']
    readonly_fields = ['submission_time', 'subtotal']
    ordering = ['-submission_time']
    
    fieldsets = (
        ('Personal Information', {
            'fields': ('first_name', 'last_name', 'email', 'phone_number')
        }),
        ('Event Details', {
            'fields': ('events', 'subtotal')
        }),
        ('Payment', {
            'fields': ('payment_screenshot', 'payment_screenshot_thumb')
        }),
        ('Metadata', {
            'fields': ('submission_time',),
            'classes': ('collapse',)
        }),
    )

    def payment_screenshot_thumb(self, obj):
        if obj.payment_screenshot:
            return format_html('<img src="{}" width="80" height="80" style="object-fit:cover; border-radius:6px; border:1px solid #ccc;" />', obj.payment_screenshot.url)
        return "No Image"
    payment_screenshot_thumb.short_description = 'Payment Screenshot'
    payment_screenshot_thumb.allow_tags = True
    
    def events_display(self, obj):
        """Display events as a formatted string"""
        if isinstance(obj.events, list):
            return ", ".join(obj.events)
        return str(obj.events)
    events_display.short_description = 'Selected Events'
    
    def full_name(self, obj):
        """Display full name"""
        return f"{obj.first_name} {obj.last_name}"
    full_name.short_description = 'Full Name'
    
    # Add export functionality
    actions = ['export_to_excel']
    
    def export_to_excel(self, request, queryset):
        """Export selected participant registrations to Excel"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participant Registrations"
        
        # Headers
        headers = [
            'ID', 'First Name', 'Last Name', 'Email', 'Phone Number', 
            'Events', 'Subtotal (Rs)', 'Registration Date', 'Payment Screenshot'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = openpyxl.styles.Font(bold=True)

        from openpyxl.drawing.image import Image as XLImage
        import os
        from django.conf import settings

        # Data rows
        for row_num, participant in enumerate(queryset, 2):
            ws[f'A{row_num}'] = participant.id
            ws[f'B{row_num}'] = participant.first_name
            ws[f'C{row_num}'] = participant.last_name
            ws[f'D{row_num}'] = participant.email
            ws[f'E{row_num}'] = participant.phone_number
            ws[f'F{row_num}'] = participant.events_display
            ws[f'G{row_num}'] = float(participant.subtotal)
            ws[f'H{row_num}'] = participant.submission_time.strftime('%Y-%m-%d %H:%M:%S')

            # Embed image if exists
            if participant.payment_screenshot:
                img_path = participant.payment_screenshot.path
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = 80
                    img.height = 80
                    cell = f'I{row_num}'
                    ws.add_image(img, cell)
                    ws.row_dimensions[row_num].height = 60
                else:
                    ws[f'I{row_num}'] = participant.payment_screenshot.url
            else:
                ws[f'I{row_num}'] = ''
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="participant_registrations.xlsx"'
        
        wb.save(response)
        return response
    
    export_to_excel.short_description = "Export selected participants to Excel"


# 34th Sammelan Payment Form Admin
@admin.register(SammelanPaymentForm)
class SammelanPaymentFormAdmin(admin.ModelAdmin):
    list_display = ['name', 'date_of_birth', 'mobile_number', 'marital_status', 'payment_screenshot_preview', 'created_at']
    search_fields = ['name', 'mobile_number', 'marital_status']
    list_filter = ['created_at']
    readonly_fields = ['created_at', 'payment_screenshot_preview', 'qr_code_preview']
    fieldsets = (
        ('Personal Information', {
            'fields': ('name', 'date_of_birth', 'mobile_number', 'marital_status')
        }),
        ('Payment Details', {
            'fields': ('qr_code_image', 'qr_code_preview', 'payment_screenshot', 'payment_screenshot_preview')
        }),
        ('Metadata', {
            'fields': ('created_at',)
        }),
    )
    
    actions = ['export_sammelan_payment_to_excel']
    
    def payment_screenshot_preview(self, obj):
        if obj.payment_screenshot:
            return format_html('<img src="{}" style="max-height:100px; max-width:150px; border-radius:8px;" />', obj.payment_screenshot.url)
        return "No screenshot"
    
    payment_screenshot_preview.short_description = "Payment Screenshot"
    
    def qr_code_preview(self, obj):
        if obj.qr_code_image:
            return format_html('<img src="{}" style="max-height:100px; max-width:150px; border-radius:8px;" />', obj.qr_code_image.url)
        return "No QR code"
    
    qr_code_preview.short_description = "QR Code"
    
    @admin.action(description='Export selected 34th Sammelan payments to Excel')
    def export_sammelan_payment_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        from PIL import Image as PILImage

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "34th Sammelan Payment"

        # Define headers
        headers = ['Name', 'Date of Birth', 'Mobile Number', 'Marital Status', 'Payment Screenshot', 'Submitted Date']
        ws.append(headers)

        # Set column widths
        column_widths = [25, 15, 15, 20, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        row_num = 2
        for obj in queryset:
            row = [
                obj.name,
                obj.date_of_birth,
                obj.mobile_number,
                obj.marital_status,
                obj.payment_screenshot.url if obj.payment_screenshot else '',
                obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            ]
            ws.append(row)
            row_num += 1

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=34th_sammelan_payment.xlsx'
        wb.save(response)
        return response

from .models import StageRegistration

@admin.register(StageRegistration)
class StageRegistrationAdmin(admin.ModelAdmin):
    list_display = [
        'name_of_candidate', 'gender', 'dob', 'current_city', 'whatsapp_number'
    ]
    search_fields = ['name_of_candidate', 'current_city', 'whatsapp_number']
    list_filter = ['gender', 'current_city']
    actions = ['export_selected_to_excel']

    @admin.action(description='Export selected Stage Registrations to Excel')
    def export_selected_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stage Registrations"
        headers = ['Full Name', 'Gender', 'Date Of Birth', 'Current City', 'WhatsApp Number']
        ws.append(headers)
        for obj in queryset:
            ws.append([
                obj.name_of_candidate,
                obj.gender,
                obj.dob,
                obj.current_city,
                obj.whatsapp_number
            ])
        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=stage_registrations.xlsx'
        wb.save(response)
        return response


@admin.register(DivorceSammelanForm)
class DivorceSammelanFormAdmin(admin.ModelAdmin):
    list_display = [
        'serial_number', 'name', 'gender', 'dob', 'marital', 'disability', 'tob', 'birthPlace', 'city', 'country',
        'visa', 'height', 'weight', 'education', 'educationDetail', 'occupationCat', 'occupationDetails',
        'salary', 'shani', 'hobbies', 'father', 'mother', 'fatherWp', 'motherWp', 'caste', 'gotra',
        'kuldevi', 'siblings', 'eating_habbits', 'alcohol', 'smoke', 'other_habbit', 'legal_case',
        'locChoice', 'ageGap', 'eduChoice', 'otherChoice', 'who', 'regMobile', 'resCat', 'nadi',
        'email', 'whatsapp', 'photo', 'declaration', 'submitted_at'
    ]

    def get_queryset(self, request):
        # Always order by submitted_at ascending for serial number stability
        qs = super().get_queryset(request)
        return qs.order_by('submitted_at')

    def serial_number(self, obj):
        # Get the queryset for the current changelist (with filters/search applied)
        request = getattr(self, 'admin_view_request', None)
        if request is None:
            return '-'
        queryset = self.get_queryset(request)
        # Get the list of primary keys in the current queryset
        pk_list = list(queryset.values_list('pk', flat=True))
        try:
            # Serial number is 1-based index in the queryset
            return pk_list.index(obj.pk) + 1
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial No.'
    serial_number.admin_order_field = None

    def get_changelist_instance(self, request):
        # Store request for use in serial_number
        self.admin_view_request = request
        return super().get_changelist_instance(request)
    search_fields = ['name', 'city', 'country', 'email', 'regMobile']
    list_filter = ['gender', 'marital', 'city', 'country', 'education', 'occupationCat']
    readonly_fields = ['submitted_at']
    ordering = ['-submitted_at']

    actions = [
        'export_selected_to_excel',
        'export_selected_to_excel_with_images',
        'download_images_zip',
    ]
    @admin.action(description='Download candidate images as ZIP')
    def download_images_zip(self, request, queryset):
        import io, zipfile, os
        from django.http import HttpResponse
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for obj in queryset:
                if hasattr(obj, 'photo') and obj.photo:
                    try:
                        img_path = obj.photo.path
                        if os.path.exists(img_path):
                            sr_number = self.serial_number(obj)
                            name = obj.name.replace(' ', '_') if hasattr(obj, 'name') else ''
                            dob = obj.dob.replace(' ', '_') if hasattr(obj, 'dob') else ''
                            filename = f"{sr_number}_{name}_{dob}{os.path.splitext(img_path)[1]}"
                            zip_file.write(img_path, f"divorce_sammelan_photos/{filename}")
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=divorce_sammelan_photos.zip'
        return response

    @admin.action(description='Export selected Divorce Sammelan Forms to Excel (with images)')
    def export_selected_to_excel_with_images(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from io import BytesIO
        from django.http import HttpResponse
        from PIL import Image as PILImage
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Divorce Sammelan Forms"

        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Photo', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)

        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]

        row_num = 2
        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            row = [
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, '', obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ]
            ws.append(row)
            # Embed photo if available
            if hasattr(obj, 'photo') and obj.photo:
                try:
                    img_path = obj.photo.path
                    if os.path.exists(img_path):
                        pil_img = PILImage.open(img_path)
                        img_byte_arr = BytesIO()
                        pil_img.thumbnail((100, 100))
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        img = OpenpyxlImage(img_byte_arr)
                        img.width = 80
                        img.height = 80
                        img_col = headers.index('Photo') + 1
                        img.anchor = f"{get_column_letter(img_col)}{row_num}"
                        ws.add_image(img)
                        ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass
            row_num += 1

        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=divorce_sammelan_forms_with_images.xlsx'
        wb.save(response)
        return response

    @admin.action(description='Export selected Divorce Sammelan Forms to Excel (without images)')
    def export_selected_to_excel_without_images(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Divorce Sammelan Forms"

        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)

        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]

        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            ws.append([
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ])

        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=divorce_sammelan_forms_without_images.xlsx'
        wb.save(response)
        return response

    @admin.action(description='Export selected Divorce Sammelan Forms to Excel')
    def export_selected_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Divorce Sammelan Forms"

        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)

        # Get queryset ordered by submitted_at for serial number stability
        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]

        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            ws.append([
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ])

        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=divorce_sammelan_forms.xlsx'
        wb.save(response)
        return response


from django import forms

@admin.register(BookletCampAdvBooking)
class BookletCampAdvBookingAdmin(admin.ModelAdmin):
    list_display = ('name', 'booking_date', 'phone', 'city', 'book', 'created_at')
    search_fields = ('name', 'phone', 'city', 'book')
    list_filter = ('booking_date', 'created_at', 'city', 'book')
    formfield_overrides = {
        # All CharFields as plain text
        forms.CharField: {'widget': forms.TextInput},
        # DateField as date input
        forms.DateField: {'widget': forms.DateInput(attrs={'type': 'date'})},
    }


# Bhudev Kalakaar 2026 Talent Registration Admin
@admin.register(BhudevKalakaar2026Registration)
class BhudevKalakaar2026RegistrationAdmin(admin.ModelAdmin):
    list_display = (
        'fullName',
        'gender',
        'dateOfBirth',
        'ageGroup',
        'talent',
        'city',
        'whatsappNumber',
        'terms',
        'photo_link',
        'submitted_at',
    )
    search_fields = ('fullName', 'city', 'whatsappNumber', 'talent', 'dateOfBirth')
    list_filter = ('gender', 'ageGroup', 'terms', 'submitted_at')
    readonly_fields = ('submitted_at', 'photo_link')
    fieldsets = (
        ('Personal Information', {
            'fields': ('fullName', 'gender', 'dateOfBirth', 'ageGroup')
        }),
        ('Talent Details', {
            'fields': ('talent',)
        }),
        ('Contact Information', {
            'fields': ('city', 'whatsappNumber')
        }),
        ('Submission', {
            'fields': ('photo', 'photo_link', 'terms', 'submitted_at')
        }),
    )
    
    def photo_link(self, obj):
        if obj.photo:
            return format_html('<a href="{}" target="_blank">View Image</a>', obj.photo.url)
        return "No photo uploaded"
    
    photo_link.short_description = "Photo Link"