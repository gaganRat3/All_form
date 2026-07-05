import io
import os
import zipfile

import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .models_38th_sammelan import Sammelan38Biodata


@admin.register(Sammelan38Biodata)
class Sammelan38BiodataAdmin(admin.ModelAdmin):
    list_display = (
        'serial_number', 'name', 'gender', 'dob', 'marital', 'disability', 'tob', 'birthPlace', 'city', 'country', 'visa',
        'height', 'weight', 'education', 'educationDetail', 'occupationCat', 'occupationDetails', 'salary', 'shani',
        'hobbies', 'father', 'mother', 'fatherWp', 'motherWp', 'caste', 'gotra', 'kuldevi', 'siblings',
        'eating_habbits', 'alcohol', 'smoke', 'other_habbit', 'legal_case', 'locChoice', 'ageGap', 'eduChoice',
        'otherChoice', 'who', 'regMobile', 'resCat', 'nadi', 'email', 'whatsapp', 'photo', 'declaration', 'submitted_at'
    )
    search_fields = (
        'name', 'city', 'email', 'whatsapp', 'father', 'mother', 'occupationDetails', 'caste', 'gotra', 'kuldevi',
        'siblings', 'educationDetail', 'regMobile', 'who'
    )
    list_filter = (
        'gender', 'marital', 'city', 'education', 'occupationCat', 'submitted_at', 'country', 'visa', 'resCat', 'nadi'
    )
    readonly_fields = ('submitted_at',)
    ordering = ['-submitted_at']

    actions = [
        'export_selected_to_excel',
        'export_selected_to_excel_with_images',
        'export_selected_to_excel_without_images',
        'download_images_zip',
    ]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('submitted_at')

    def serial_number(self, obj):
        request = getattr(self, 'admin_view_request', None)
        if request is None:
            return '-'
        queryset = self.get_queryset(request)
        pk_list = list(queryset.values_list('pk', flat=True))
        try:
            return pk_list.index(obj.pk) + 1
        except ValueError:
            return '-'
    serial_number.short_description = 'Serial No.'
    serial_number.admin_order_field = None

    def get_changelist_instance(self, request):
        self.admin_view_request = request
        return super().get_changelist_instance(request)

    @admin.action(description='Download candidate images as ZIP')
    def download_images_zip(self, request, queryset):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for obj in queryset:
                if hasattr(obj, 'photo') and obj.photo:
                    try:
                        img_path = obj.photo.path
                        if os.path.exists(img_path):
                            sr_number = self.serial_number(obj)
                            name = obj.name.replace(' ', '_') if hasattr(obj, 'name') else ''
                            dob = obj.dob.replace(' ', '_') if hasattr(obj, 'dob') else ''
                            filename = f"{sr_number}_{name}_{dob}{os.path.splitext(img_path)[1]}"
                            zip_file.write(img_path, f"sammelan38_photos/{filename}")
                    except Exception:
                        pass
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=38th_sammelan_photos.zip'
        return response

    @admin.action(description='Export selected 38th Sammelan Forms to Excel (with images)')
    def export_selected_to_excel_with_images(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "38th Sammelan Forms"
        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Photo', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)
        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]
        row_num = 2
        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            row = [
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, '', obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ]
            ws.append(row)
            if hasattr(obj, 'photo') and obj.photo:
                try:
                    img_path = obj.photo.path
                    if os.path.exists(img_path):
                        pil_img = PILImage.open(img_path)
                        img_byte_arr = io.BytesIO()
                        pil_img.thumbnail((100, 100))
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        img = OpenpyxlImage(img_byte_arr)
                        img.width = 80
                        img.height = 80
                        img_col = headers.index('Photo') + 1
                        img.anchor = f"{get_column_letter(img_col)}{row_num}"
                        ws.add_image(img)
                        ws.row_dimensions[row_num].height = 60
                except Exception:
                    pass
            row_num += 1
        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=38th_sammelan_forms_with_images.xlsx'
        wb.save(response)
        return response

    @admin.action(description='Export selected 38th Sammelan Forms to Excel (without images)')
    def export_selected_to_excel_without_images(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "38th Sammelan Forms"
        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)
        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]
        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            ws.append([
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ])
        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=38th_sammelan_forms_without_images.xlsx'
        wb.save(response)
        return response

    @admin.action(description='Export selected 38th Sammelan Forms to Excel')
    def export_selected_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "38th Sammelan Forms"
        headers = [
            'Serial No.', 'Name', 'Gender', 'DOB', 'Marital Status', 'Disability', 'Time of Birth', 'Birth Place', 'City', 'Country',
            'Visa', 'Height', 'Weight', 'Education', 'Education Detail', 'Occupation Category', 'Occupation Details',
            'Salary', 'Shani', 'Hobbies', 'Father', 'Mother', 'Father WhatsApp', 'Mother WhatsApp', 'Caste', 'Gotra',
            'Kuldevi', 'Siblings', 'Eating Habits', 'Alcohol', 'Smoke', 'Other Habit', 'Legal Case',
            'Location Choice', 'Age Gap', 'Education Choice', 'Other Choice', 'Who', 'Registered Mobile', 'Residence Category', 'Nadi',
            'Email', 'WhatsApp', 'Declaration', 'Submitted At'
        ]
        ws.append(headers)
        export_queryset = self.get_queryset(request)
        pk_list = list(export_queryset.values_list('pk', flat=True))
        selected_ids = set(queryset.values_list('pk', flat=True))
        export_objs = [obj for obj in export_queryset if obj.pk in selected_ids]
        for obj in export_objs:
            try:
                serial_no = pk_list.index(obj.pk) + 1
            except ValueError:
                serial_no = '-'
            ws.append([
                serial_no, obj.name, obj.gender, obj.dob, obj.marital, obj.disability, obj.tob, obj.birthPlace, obj.city, obj.country,
                obj.visa, obj.height, obj.weight, obj.education, obj.educationDetail, obj.occupationCat, obj.occupationDetails,
                obj.salary, obj.shani, obj.hobbies, obj.father, obj.mother, obj.fatherWp, obj.motherWp, obj.caste, obj.gotra,
                obj.kuldevi, obj.siblings, obj.eating_habbits, obj.alcohol, obj.smoke, obj.other_habbit, obj.legal_case,
                obj.locChoice, obj.ageGap, obj.eduChoice, obj.otherChoice, obj.who, obj.regMobile, obj.resCat, obj.nadi,
                obj.email, obj.whatsapp, obj.declaration, obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if obj.submitted_at else ''
            ])
        for i, column in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=38th_sammelan_forms.xlsx'
        wb.save(response)
        return response
